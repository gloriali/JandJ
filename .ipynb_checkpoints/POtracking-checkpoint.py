# combine PO files and shipping info into PO tracking file
season = "2023FW"

# ------ write header for new Shipment Tracking.xlsx file ------
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
thin = Side(border_style = "thin", color = "000000")
thick = Side(border_style = "thick", color = "0033CCCC")
POtrack = Workbook()
POtrack1 = POtrack.active
POtrack1.title = season
POtrack1.merge_cells("A1:J3")
POtrack1["A1"] = season + " China to Global Shipment Tracking"
POtrack1["A1"].font = Font(size = 22, bold = True)
POtrack1["A1"].alignment = center = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
POtrack1.merge_cells("K1:L1")
POtrack1["K1"] = "Country/region"
POtrack1["K1"].alignment = center
POtrack1.merge_cells("K2:L2")
POtrack1["K2"] = "Container"
POtrack1["K2"].alignment = center
POtrack1.merge_cells("K3:L3")
POtrack1["K3"] = "Shipment Details"
POtrack1["K3"].alignment = center
POtrack1.merge_cells("M1:O1")
POtrack1["M1"] = "CA/US"
POtrack1["M1"].font = header = Font(size = 16, bold = True)
POtrack1["M1"].alignment = center
POtrack1["M1"].fill = PatternFill("solid", fgColor = "00008000")
POtrack1.merge_cells("P1:R1")
POtrack1["P1"] = "UK"
POtrack1["P1"].font = header 
POtrack1["P1"].alignment = center
POtrack1["P1"].fill = PatternFill("solid", fgColor = "003366FF")
POtrack1.merge_cells("S1:U1")
POtrack1["S1"] = "DE"
POtrack1["S1"].font = header 
POtrack1["S1"].alignment = center
POtrack1["S1"].fill = PatternFill("solid", fgColor = "00FF9900")
POtrack1.merge_cells("V1:X1")
POtrack1["V1"] = "CN Initial Reserve"
POtrack1["V1"].font = header 
POtrack1["V1"].alignment = center
POtrack1["V1"].fill = PatternFill("solid", fgColor = "00FF0000")
POtrack1.merge_cells("Y1:AA1")
POtrack1["Y1"] = "Wholesaler"
POtrack1["Y1"].font = header 
POtrack1["Y1"].alignment = center
POtrack1["Y1"].fill = PatternFill("solid", fgColor = "00339966")
head = ["PO#",	"CAT", "Season", "Factory", "Total PO in Pcs", "Name", "SKU", "Seasons", "Colour/Graphic(EN)", "Colour/Graphic(CN)", "Remain. in Pcs", "Remain. in %"] + ["PO Plan", "Remain. in Pcs", "Remain. in %"]*5
for col in range(1, 28):
     POtrack1.cell(4, col).value = head[col-1]
     POtrack1.cell(4, col).font = Font(size = 11, bold = True)
     POtrack1.cell(4, col).alignment = center
     POtrack1.cell(4, col).fill = PatternFill("solid", fgColor = "00FFCC00")
     POtrack1.cell(4, col).border = Border(top = thin, bottom = thin, left = thin, right = thin)
POtrack1.row_dimensions[1].border = Border(bottom = thin)
POtrack1.row_dimensions[2].border = Border(bottom = thin)
POtrack1.row_dimensions[3].border = Border(bottom = thin)
POtrack1.column_dimensions['J'].border = Border(right = thin)
POtrack1.column_dimensions['M'].border = Border(left = thick)
POtrack1.column_dimensions['P'].border = Border(left = thick)
POtrack1.column_dimensions['S'].border = Border(left = thick)
POtrack1.column_dimensions['V'].border = Border(left = thick)
POtrack1.column_dimensions['Y'].border = Border(left = thick)
POtrack.save("../PO/" + season + " China to Global Shipment Tracking.xlsx")

# ----------- input PO -------------
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
masterSKUfile = load_workbook(filename = "../PO/1-MasterSKU-All-Product-2023-11-07.xlsx").active
masterSKU = DataFrame(masterSKUfile.values)
POtrack = load_workbook(filename = "../PO/" + season + " China to Global Shipment Tracking.xlsx")
POtrack1 = POtrack[season]

