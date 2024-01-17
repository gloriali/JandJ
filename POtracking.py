# combine PO files and shipping info into PO tracking file
season = "2024SS"

# Function: Search value in dataframe, and find the next n non-empty cell
def next_non_empty(df, value, dim, n):
  mask = df.applymap(lambda x: str(value) in str(x)).to_numpy()
  if mask.any():
    r, c = np.argwhere(mask)[0]
  else:
    return df.shape[0]+1, df.shape[1]+1
  if dim == "row":
    for row in range(r+n, df.shape[0]):
      if df.iloc[row, c] is not (np.nan and None): 
        return row, c
        break
  if dim == "col": 
    for col in range(c+n, df.shape[1]):
      if df.iloc[r, col] is not (np.nan and None):
        return r, col
        break

# ------ write header for new Shipment Tracking.xlsx file ------
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
if os.path.exists("../PO/" + season + " China to Global Shipment Tracking.xlsx"):
  os.remove("../PO/" + season + " China to Global Shipment Tracking.xlsx")
thin = Side(border_style = "thin", color = "000000")
thick = Side(border_style = "thick", color = "0033CCCC")
POtrack = Workbook()
POtrack1 = POtrack.active
POtrack1.title = season
POtrack1.merge_cells("A1:J3")
POtrack1["A1"] = season + " China to Global Shipment Tracking"
POtrack1["A1"].font = Font(size = 22, bold = True)
POtrack1["A1"].alignment = center = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
POtrack1.merge_cells("K1:L1")
POtrack1["K1"] = "Country/region"
POtrack1["K1"].alignment = center
POtrack1.merge_cells("K2:L2")
POtrack1["K2"] = "Container"
POtrack1["K2"].alignment = center
POtrack1.merge_cells("K3:L3")
POtrack1["K3"] = "Shipment Details"
POtrack1["K3"].alignment = center
POtrack1.column_dimensions["M"].fill = PatternFill("solid", fgColor = "00339966")
POtrack1.column_dimensions["M"].width = 1
POtrack1.merge_cells("N1:P1")
POtrack1["N1"] = "CA/US"
POtrack1["N1"].font = header = Font(size = 16, bold = True)
POtrack1["N1"].alignment = center
POtrack1["N1"].fill = PatternFill("solid", fgColor = "00008000")
POtrack1.column_dimensions["Q"].fill = PatternFill("solid", fgColor = "00339966")
POtrack1.column_dimensions["Q"].width = 1
POtrack1.merge_cells("R1:T1")
POtrack1["R1"] = "UK"
POtrack1["R1"].font = header 
POtrack1["R1"].alignment = center
POtrack1["R1"].fill = PatternFill("solid", fgColor = "003366FF")
POtrack1.column_dimensions["U"].fill = PatternFill("solid", fgColor = "00339966")
POtrack1.column_dimensions["U"].width = 1
POtrack1.merge_cells("V1:X1")
POtrack1["V1"] = "DE"
POtrack1["V1"].font = header 
POtrack1["V1"].alignment = center
POtrack1["V1"].fill = PatternFill("solid", fgColor = "00FF9900")
POtrack1.column_dimensions["Y"].fill = PatternFill("solid", fgColor = "00339966")
POtrack1.column_dimensions["Y"].width = 1
POtrack1.merge_cells("Z1:AB1")
POtrack1["Z1"] = "CN Initial Reserve"
POtrack1["Z1"].font = header 
POtrack1["Z1"].alignment = center
POtrack1["Z1"].fill = PatternFill("solid", fgColor = "00FF0000")
POtrack1.column_dimensions["AC"].fill = PatternFill("solid", fgColor = "00339966")
POtrack1.column_dimensions["AC"].width = 1
POtrack1.merge_cells("AD1:AF1")
POtrack1["AD1"] = "Wholesaler"
POtrack1["AD1"].font = header 
POtrack1["AD1"].alignment = center
POtrack1["AD1"].fill = PatternFill("solid", fgColor = "00339966")
head = ["PO#", "CAT", "Season", "Factory", "Total PO in Pcs", "Name", "SKU", "Seasons", "Colour/Graphic(EN)", "Colour/Graphic(CN)", "Remain. in Pcs", "Remain. in %"] + ["", "PO Plan", "Remain. in Pcs", "Remain. in %"]*5
for col in range(1, 33):
  POtrack1.cell(4, col).value = head[col-1]
  POtrack1.cell(4, col).font = Font(size = 11, bold = True)
  POtrack1.cell(4, col).alignment = center
  POtrack1.cell(4, col).fill = PatternFill("solid", fgColor = "00FFCC00")
  POtrack1.cell(4, col).border = Border(top = thin, bottom = thin, left = thin, right = thin)
POtrack1.row_dimensions[1].border = Border(bottom = thin)
POtrack1.row_dimensions[2].border = Border(bottom = thin)
POtrack1.row_dimensions[3].border = Border(bottom = thin)
POtrack1.freeze_panes = POtrack1['N5']
POtrack.save("../PO/" + season + " China to Global Shipment Tracking.xlsx")

# ----------- input PO -------------
import pandas as pd
import numpy as np
import os
import re
from glob import glob
from openpyxl import Workbook
from openpyxl import load_workbook

## Factory code and master SKU
factory = pd.read_excel("../../TWK 2020 share/1-MasterSKU_CatPrintsFactory.xlsx", sheet_name = 3, skiprows = 2, engine = "openpyxl")
factory.drop(columns = ["Unnamed: 7", "Unnamed: 8", "Unnamed: 9", "Unnamed: 10"], inplace = True) 
factory.rename(columns = {"Unnamed: 2": "JSMY"}, inplace=True)
factory["id"] = factory.index
factory = pd.melt(factory, id_vars = "id")
factory = factory.dropna(axis = 0, how = "any")
factory = pd.Series(factory.variable.values, index = factory.value).to_dict()
masterSKU = pd.read_excel(glob(os.path.join("../../TWK 2020 share/1-MasterSKU-All-Product-" + "*.xlsx"))[-1], sheet_name = 0, skiprows = 1, header = 2, usecols = "B:M", engine = "openpyxl")
masterSKU.MSKU = [x.upper() for x in masterSKU.MSKU]
masterSKU = pd.Series(masterSKU.Seasons.values, index = masterSKU.MSKU).to_dict()

## input PO files
POn = ["P" + str(n) for n in range(257, 288)]
PO_checklist = fac_checklist = cat_checklist = []
for i in POn:
  print(i)
  POtrack1 = load_workbook("../PO/" + season + " China to Global Shipment Tracking.xlsx")[season]
  POtrack_df = pd.DataFrame(POtrack1.values)
  PO_r = POtrack_df.shape[0]
  (Detail_r, Detail_c) = POtrack_df.where(POtrack_df.eq("Total PO in Pcs")).stack().index.tolist()[0]
  (T_r, T_c) = POtrack_df.where(POtrack_df.eq("Country/region")).stack().index.tolist()[0]
  (CA_r, CA_c) = POtrack_df.where(POtrack_df.eq("CA/US")).stack().index.tolist()[0]
  (UK_r, UK_c) = POtrack_df.where(POtrack_df.eq("UK")).stack().index.tolist()[0]
  (DE_r, DE_c) = POtrack_df.where(POtrack_df.eq("DE")).stack().index.tolist()[0]
  (CN_r, CN_c) = POtrack_df.where(POtrack_df.eq("CN Initial Reserve")).stack().index.tolist()[0]    
  (WS_r, WS_c) = POtrack_df.where(POtrack_df.eq("Wholesaler")).stack().index.tolist()[0]
  file = [y for x in os.walk("../PO/order/" + season + "/") for y in glob(os.path.join(x[0], i + "*.xlsx"))]
  print(file)
  if file == []: 
    continue
  name = cat = fac = ''
  TotalPO = SKU = Seasons = color_EN = color_CN = Remain_Pc = CA_POplan = CA_Remain_Pc = UK_POplan = UK_Remain_Pc = DE_POplan = DE_Remain_Pc = CN_POplan = CN_Remain_Pc = WS_POplan = WS_Remain_Pc = []
  for f in file:
    name = name + re.sub('\\\.*', '', re.sub('.*/'+ i +' - ', '', f))
    print(name)
    PO = pd.read_excel(f, sheet_name = 0, engine = "openpyxl", header = None)
    total = PO.iloc[next_non_empty(PO, "OrderTotal", "col", 1)]
    destination = [x.replace(' ', '').replace('\n', '').replace('：', ':') for x in PO[PO.applymap(lambda x: '总数量' in str(x))].stack().tolist()]
    ws = [x for x in destination if x not in ['订单总数量OrderTotal:', '订单总数量', '加拿大总数量', '英国总数量', '德国总数量', '中国总数量', '总数量TotalQty:']]
    try: (CA_PO_r, CA_PO_c) = PO.where(PO.eq('加拿大总数量')).stack().index.tolist()[0]
    except (KeyError, IndexError): (CA_PO_r, CA_PO_c) = (PO.shape[0]+1, PO.shape[1]+1)
    try: (UK_PO_r, UK_PO_c) = PO.where(PO.eq('英国总数量')).stack().index.tolist()[0]
    except (KeyError, IndexError): (UK_PO_r, UK_PO_c) = (PO.shape[0]+1, PO.shape[1]+1)
    try: (DE_PO_r, DE_PO_c) = PO.where(PO.eq('德国总数量')).stack().index.tolist()[0]
    except (KeyError, IndexError): (DE_PO_r, DE_PO_c) = (PO.shape[0]+1, PO.shape[1]+1)
    try: (CN_PO_r, CN_PO_c) = PO.where(PO.eq('中国总数量')).stack().index.tolist()[0]
    except (KeyError, IndexError): (CN_PO_r, CN_PO_c) = (PO.shape[0]+1, PO.shape[1]+1)
    try: (WS_PO_r, WS_PO_c) = PO.where(PO.eq(ws[0])).stack().index.tolist()[0]
    except (KeyError, IndexError): (WS_PO_r, WS_PO_c) = (PO.shape[0]+1, PO.shape[1]+1)
    (SKU_r, SKU_c) = PO.where(PO.eq('SKU')).stack().index.tolist()[0]
    PO_detail = pd.DataFrame(data = PO.iloc[SKU_r+2:PO.shape[0], SKU_c:PO.shape[1]])
    PO_detail.columns = PO.iloc[SKU_r-1, SKU_c:PO.shape[1]].replace('\n', '', regex = True).replace('\s', '', regex = True).replace('花型/颜色(EN)', '英文品名').replace('花型/颜色(CN)', '中文品名')
    PO_detail['产品编号'] = PO_detail['产品编号'].str.strip()
    PO_detail = PO_detail[PO_detail['产品编号'].notna()]
    PO_detail = PO_detail[PO_detail['产品编号'].str.contains('-')]
    if total != PO_detail['订单总数量'].sum(): 
      print('ERROR: Total in Pcs does not match.', total, PO_detail['订单总数量'].sum())
    
    cat_list = list(dict.fromkeys(PO_detail['产品编号'].replace('-.*', '', regex = True).tolist()))
    fac_list = list(dict.fromkeys([factory[x] for x in cat_list]))
    cat = cat + ' '.join(cat_list)
    fac = fac + ' '.join(fac_list)
    TotalPO = TotalPO + PO_detail['订单总数量'].tolist()
    SKU = SKU + PO_detail['产品编号'].tolist()
    Seasons = Seasons + [masterSKU.get(x.upper(), "NaN") for x in PO_detail['产品编号']] 
    color_EN = color_EN + PO_detail['英文品名'].tolist()
    color_CN = color_CN + PO_detail['中文品名'].tolist()
    Remain_Pc = Remain_Pc + PO_detail['订单总数量'].tolist()
    try: 
      CA_POplan = CA_POplan + PO_detail.iloc[:, CA_PO_c-SKU_c].tolist()
      CA_Remain_Pc = CA_Remain_Pc + PO_detail.iloc[:, CA_PO_c-SKU_c].tolist()
    except (KeyError, IndexError): 
      print('No CA/US')
      CA_POplan = CA_POplan + [' ' * PO_detail.shape[0]]
      CA_Remain_Pc = CA_Remain_Pc + [' ' * PO_detail.shape[0]]
    
    try: 
      UK_POplan = UK_POplan + PO_detail.iloc[:, UK_PO_c-SKU_c].tolist()
      UK_Remain_Pc = UK_Remain_Pc + PO_detail.iloc[:, UK_PO_c-SKU_c].tolist()
    except (KeyError, IndexError): 
      print('No UK')
      UK_POplan = UK_POplan + [' ' * PO_detail.shape[0]]
      UK_Remain_Pc = UK_Remain_Pc + [' ' * PO_detail.shape[0]]
    
    try: 
      DE_POplan = DE_POplan + PO_detail.iloc[:, DE_PO_c-SKU_c].tolist()
      DE_Remain_Pc = DE_Remain_Pc + PO_detail.iloc[:, DE_PO_c-SKU_c].tolist()
    except (KeyError, IndexError): 
      print('No DE')
      DE_POplan = DE_POplan + [' ' * PO_detail.shape[0]]
      DE_Remain_Pc = DE_Remain_Pc + [' ' * PO_detail.shape[0]]
    
    try: 
      CN_POplan = CN_POplan + PO_detail.iloc[:, CN_PO_c-SKU_c].tolist()
      CN_Remain_Pc = CN_Remain_Pc + PO_detail.iloc[:, CN_PO_c-SKU_c].tolist()
    except (KeyError, IndexError): 
      print('No CN')
      CN_POplan = CN_POplan + [' ' * PO_detail.shape[0]]
      CN_Remain_Pc = CN_Remain_Pc + [' ' * PO_detail.shape[0]]
    
    try: 
      WS_POplan = WS_POplan + PO_detail.iloc[:, WS_PO_c-SKU_c].tolist()
      WS_Remain_Pc = WS_Remain_Pc + PO_detail.iloc[:, WS_PO_c-SKU_c].tolist()
    except (KeyError, IndexError): 
      print('No Wholesaler')
      WS_POplan = WS_POplan + [' ' * PO_detail.shape[0]]
      WS_Remain_Pc = WS_Remain_Pc + [' ' * PO_detail.shape[0]]
    
  detail = pd.DataFrame({'TotalPO': TotalPO, 'Name': '', 'SKU': SKU, 'seasons': Seasons, 'color_EN': color_EN, 'color_CN': color_CN, 'Remain_Pc': Remain_Pc, 'Remain_percent': format(1,'.1%')})
  subtotal = pd.DataFrame({'PO': [i], 'CAT': [cat], 'Season': [season], 'Factory': [fac], 'Total': [detail['TotalPO'].sum()], 'Name': [name], 'SKU': '', 'seasons': '', 'color_EN': '', 'color_CN': '', 'Remain_Pc': [detail['Remain_Pc'].sum()], 'Remain_percent': format(1,'.1%')})
  with pd.ExcelWriter("../PO/" + season + " China to Global Shipment Tracking.xlsx", if_sheet_exists = 'overlay', mode = 'a') as writer:
    subtotal.to_excel(writer, sheet_name = season, startrow = PO_r, startcol = 0, header = None, index = False)
    detail.to_excel(writer, sheet_name = season, startrow = PO_r+1, startcol = Detail_c, header = None, index = False)
  
  CA_detail = pd.DataFrame({'POplan': CA_POplan, 'Remain_Pc': CA_Remain_Pc, 'Remain_percent': '' if all(s == '' or str(s).isspace() for s in CA_POplan) else format(1,'.1%')})
  CA_subtotal = pd.DataFrame({'POplan': [CA_detail['POplan'].sum()], 'Remain_Pc': [CA_detail['Remain_Pc'].sum()], 'Remain_percent': '' if all(s == '' or str(s).isspace() for s in CA_POplan) else format(1,'.1%')})
  with pd.ExcelWriter("../PO/" + season + " China to Global Shipment Tracking.xlsx", if_sheet_exists = 'overlay', mode = 'a') as writer:
    CA_subtotal.to_excel(writer, sheet_name = season, startrow = PO_r, startcol = CA_c, header = None, index = False)
    CA_detail.to_excel(writer, sheet_name = season, startrow = PO_r+1, startcol = CA_c, header = None, index = False)
  
  UK_detail = pd.DataFrame({'POplan': UK_POplan, 'Remain_Pc': UK_Remain_Pc, 'Remain_percent': '' if all(s == '' or str(s).isspace() for s in UK_POplan) else format(1,'.1%')})
  UK_subtotal = pd.DataFrame({'POplan': [UK_detail['POplan'].sum()], 'Remain_Pc': [UK_detail['Remain_Pc'].sum()], 'Remain_percent': '' if all(s == '' or str(s).isspace() for s in UK_POplan) else format(1,'.1%')})
  with pd.ExcelWriter("../PO/" + season + " China to Global Shipment Tracking.xlsx", if_sheet_exists = 'overlay', mode = 'a') as writer:
    UK_subtotal.to_excel(writer, sheet_name = season, startrow = PO_r, startcol = UK_c, header = None, index = False)
    UK_detail.to_excel(writer, sheet_name = season, startrow = PO_r+1, startcol = UK_c, header = None, index = False)
  
  DE_detail = pd.DataFrame({'POplan': DE_POplan, 'Remain_Pc': DE_Remain_Pc, 'Remain_percent': '' if all(s == '' or str(s).isspace() for s in DE_POplan) else format(1,'.1%')})
  DE_subtotal = pd.DataFrame({'POplan': [DE_detail['POplan'].sum()], 'Remain_Pc': [DE_detail['Remain_Pc'].sum()], 'Remain_percent': '' if all(s == '' or str(s).isspace() for s in DE_POplan) else format(1,'.1%')})
  with pd.ExcelWriter("../PO/" + season + " China to Global Shipment Tracking.xlsx", if_sheet_exists = 'overlay', mode = 'a') as writer:
    DE_subtotal.to_excel(writer, sheet_name = season, startrow = PO_r, startcol = DE_c, header = None, index = False)
    DE_detail.to_excel(writer, sheet_name = season, startrow = PO_r+1, startcol = DE_c, header = None, index = False)
  
  CN_detail = pd.DataFrame({'POplan': CN_POplan, 'Remain_Pc': CN_Remain_Pc, 'Remain_percent': '' if all(s == '' or str(s).isspace() for s in CN_POplan) else format(1,'.1%')})
  CN_subtotal = pd.DataFrame({'POplan': [CN_detail['POplan'].sum()], 'Remain_Pc': [CN_detail['Remain_Pc'].sum()], 'Remain_percent': '' if all(s == '' or str(s).isspace() for s in CN_POplan) else format(1,'.1%')})
  with pd.ExcelWriter("../PO/" + season + " China to Global Shipment Tracking.xlsx", if_sheet_exists = 'overlay', mode = 'a') as writer:
    CN_subtotal.to_excel(writer, sheet_name = season, startrow = PO_r, startcol = CN_c, header = None, index = False)
    CN_detail.to_excel(writer, sheet_name = season, startrow = PO_r+1, startcol = CN_c, header = None, index = False)
  
  WS_detail = pd.DataFrame({'POplan': WS_POplan, 'Remain_Pc': WS_Remain_Pc, 'Remain_percent': '' if all(s == '' or str(s).isspace() for s in WS_POplan) else format(1,'.1%')})
  WS_subtotal = pd.DataFrame({'POplan': [WS_detail['POplan'].sum()], 'Remain_Pc': [WS_detail['Remain_Pc'].sum()], 'Remain_percent': '' if all(s == '' or str(s).isspace() for s in WS_POplan) else format(1,'.1%')})
  with pd.ExcelWriter("../PO/" + season + " China to Global Shipment Tracking.xlsx", if_sheet_exists = 'overlay', mode = 'a') as writer:
    WS_subtotal.to_excel(writer, sheet_name = season, startrow = PO_r, startcol = WS_c, header = None, index = False)
    WS_detail.to_excel(writer, sheet_name = season, startrow = PO_r+1, startcol = WS_c, header = None, index = False)
  
  POtrack = load_workbook("../PO/" + season + " China to Global Shipment Tracking.xlsx")
  POtrack1 = POtrack[season]
  POtrack1.sheet_properties.outlinePr.summaryBelow = False
  POtrack1.row_dimensions.group(PO_r+2, PO_r+len(SKU)+1, outline_level = 1)
  POtrack.save("../PO/" + season + " China to Global Shipment Tracking.xlsx")
  
  PO_checklist = PO_checklist + [i]
  fac_checklist = fac_checklist + [fac]
  cat_checklist = cat_checklist + [cat]

checklist = pd.DataFrame({'Category': cat_checklist, 'Factory': fac_checklist, 'PO': PO_checklist})
with pd.ExcelWriter("../PO/" + season + "_PreProduction_checklist.xlsx", mode = 'w') as writer:
  checklist.to_excel(writer, sheet_name = 'checklist', header = None, index = False)

POtrack = load_workbook("../PO/" + season + " China to Global Shipment Tracking.xlsx")
POtrack1 = POtrack[season]
POtrack_df = pd.DataFrame(POtrack1.values)
for i in range(len(POtrack_df)):
  if POtrack1.cell(i + 1, 5).value == 0:
    POtrack1.cell(i + 1, 12).value = format(0, ".1%")
    POtrack1.cell(i + 1, 12).font = Font(color = "0000FF00")
  if POtrack1.cell(i + 1, 14).value == 0:
    POtrack1.cell(i + 1, 16).value = format(0, ".1%")
    POtrack1.cell(i + 1, 16).font = Font(color = "0000FF00")
  if POtrack1.cell(i + 1, 18).value == 0:
    POtrack1.cell(i + 1, 20).value = format(0, ".1%")
    POtrack1.cell(i + 1, 20).font = Font(color = "0000FF00")
  if POtrack1.cell(i + 1, 22).value == 0:
    POtrack1.cell(i + 1, 24).value = format(0, ".1%")
    POtrack1.cell(i + 1, 24).font = Font(color = "0000FF00")
  if POtrack1.cell(i + 1, 26).value == 0:
    POtrack1.cell(i + 1, 28).value = format(0, ".1%")
    POtrack1.cell(i + 1, 28).font = Font(color = "0000FF00")
  if POtrack1.cell(i + 1, 30).value == 0:
    POtrack1.cell(i + 1, 32).value = format(0, ".1%")
    POtrack1.cell(i + 1, 32).font = Font(color = "0000FF00")

POtrack.save("../PO/" + season + " China to Global Shipment Tracking.xlsx")

# ----------- input shipment: CA - Surrey -------------
import pandas as pd
import numpy as np
import os
import re
from glob import glob
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

# !! insert column for shipment in tracking file in Excel first !!
season = "2024SS"
ship = "S1"
total_c = 5 # column number for Total PO in Pcs
remaint_c = 11 # column number for total remain in Pcs
remainr_c = 15 # column number for regional remain in Pcs
insert_c = 17 # column number for shipment in tracking file
thin = Side(border_style = "thin", color = "000000")
thick = Side(border_style = "thick", color = "0033CCCC")
center = Alignment(horizontal = "center", vertical = "center", wrap_text = True)
ship_f = glob(os.path.join("../PO/shipment/*" + ship + "*/*" +  ship + "*.xlsx"))[0]
shipment = load_workbook(ship_f)
sheets = shipment.get_sheet_names()
sku_all = qty_all = PO_all = []
for i in range(1, len(sheets)):
  print(sheets[i])
  shipment_i = pd.read_excel(ship_f, sheet_name = sheets[i], engine = "openpyxl", header = 0, usecols = "A:G", skiprows = 4).dropna(subset = ["SKU", "PO #"])
  sku_all = sku_all + shipment_i["SKU"].to_list()
  qty_all = qty_all + shipment_i["QTY SHIPPED"].to_list()
  PO_all = PO_all + shipment_i["PO #"].to_list()

shipment_all = pd.DataFrame({"PO": PO_all, "SKU": sku_all, "QTY": qty_all}).groupby(["PO", "SKU"]).sum().reset_index()
shipment_PO = shipment_all.groupby(["PO"]).sum().reset_index()

POtrack = load_workbook("../PO/" + season + " China to Global Shipment Tracking.xlsx")
POtrack1 = POtrack[season]
POtrack_df = pd.DataFrame(POtrack1.values)
POtrack1.cell(1, insert_c).value = ship
POtrack1.cell(1, insert_c).alignment = center
POtrack1.cell(4, insert_c).value = "SKU & Qty (Pcs)"
POtrack1.cell(4, insert_c).font = Font(size = 11, bold = True)
POtrack1.cell(4, insert_c).alignment = center
POtrack1.cell(4, insert_c).fill = PatternFill("solid", fgColor = "00FFCC00")
POtrack1.cell(4, insert_c).border = Border(top = thin, bottom = thin, left = thin, right = thin)
for i in range(len(shipment_PO)):
  print(shipment_PO.loc[i, "PO"])
  mask = POtrack_df.applymap(lambda x: str(shipment_PO.loc[i, "PO"]) in str(x)).to_numpy()
  r, c = np.argwhere(mask)[0]
  POtrack1.cell(r + 1, insert_c).value = shipment_PO.loc[i, "QTY"]
  POtrack1.cell(r + 1, remaint_c).value = POtrack1.cell(r + 1, remaint_c).value - shipment_PO.loc[i, "QTY"]
  POtrack1.cell(r + 1, remaint_c + 1).value = format(POtrack1.cell(r + 1, remaint_c).value / POtrack1.cell(r + 1, total_c).value, ".1%")
  if float(POtrack1.cell(r + 1, remaint_c + 1).value.strip("%")) > 1 or float(POtrack1.cell(r + 1, remaint_c + 1).value.strip("%")) < -10: 
    POtrack1.cell(r + 1, remaint_c + 1).font = Font(color = "00FF0000")
  else:
    POtrack1.cell(r + 1, remaint_c + 1).font = Font(color = "0000FF00")
  try: 
    POtrack1.cell(r + 1, remainr_c).value = POtrack1.cell(r + 1, remainr_c).value - shipment_PO.loc[i, "QTY"]
    POtrack1.cell(r + 1, remainr_c + 1).value = format(POtrack1.cell(r + 1, remainr_c).value / POtrack1.cell(r + 1, remainr_c - 1).value, ".1%")
    if float(POtrack1.cell(r + 1, remainr_c + 1).value.strip("%")) > 1 or float(POtrack1.cell(r + 1, remainr_c + 1).value.strip("%")) < -10: 
      POtrack1.cell(r + 1, remainr_c + 1).font = Font(color = "00FF0000")
    else:
      POtrack1.cell(r + 1, remainr_c + 1).font = Font(color = "0000FF00")
  except: print("ERROR")

for i in range(len(shipment_all)):
  print(shipment_all.loc[i, "PO"], shipment_all.loc[i, "SKU"])
  mask = POtrack_df.applymap(lambda x: str(shipment_all.loc[i, "PO"]) in str(x)).to_numpy()
  r_1, c = np.argwhere(mask)[0]
  r_2, c = next_non_empty(POtrack_df, shipment_all.loc[i, "PO"], "row", 1)
  mask = POtrack_df.applymap(lambda x: str(shipment_all.loc[i, "SKU"]) in str(x)).to_numpy()
  for j in range(len(np.argwhere(mask))): 
    r, c = np.argwhere(mask)[j]
    if r > r_1 and r < r_2: 
      break
  POtrack1.cell(r + 1, insert_c).value = shipment_all.loc[i, "QTY"]
  POtrack1.cell(r + 1, remaint_c).value = POtrack1.cell(r + 1, remaint_c).value - shipment_all.loc[i, "QTY"]
  POtrack1.cell(r + 1, remaint_c + 1).value = format(POtrack1.cell(r + 1, remaint_c).value / POtrack1.cell(r + 1, total_c).value, ".1%")
  if float(POtrack1.cell(r + 1, remaint_c + 1).value.strip("%")) > 1 or float(POtrack1.cell(r + 1, remaint_c + 1).value.strip("%")) < -1: 
    POtrack1.cell(r + 1, remaint_c + 1).font = Font(color = "00FF0000")
  else:
    POtrack1.cell(r + 1, remaint_c + 1).font = Font(color = "0000FF00")
  try: 
    POtrack1.cell(r + 1, remainr_c).value = POtrack1.cell(r + 1, remainr_c).value - shipment_all.loc[i, "QTY"]
    POtrack1.cell(r + 1, remainr_c + 1).value = format(POtrack1.cell(r + 1, remainr_c).value / POtrack1.cell(r + 1, remainr_c - 1).value, ".1%")
    if float(POtrack1.cell(r + 1, remainr_c + 1).value.strip("%")) > 1 or float(POtrack1.cell(r + 1, remainr_c + 1).value.strip("%")) < -1: 
      POtrack1.cell(r + 1, remainr_c + 1).font = Font(color = "00FF0000")
    else:
      POtrack1.cell(r + 1, remainr_c + 1).font = Font(color = "0000FF00")
  except: print("ERROR")

POtrack.save("../PO/" + season + " China to Global Shipment Tracking1.xlsx")
